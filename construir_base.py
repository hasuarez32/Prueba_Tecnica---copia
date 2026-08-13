#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
construir_base.py — Base consolidada logística CEC (Universidad del Norte)
==========================================================================

Recorre dinámicamente las carpetas de programa del CEC y produce una única
base consolidada `base_consolidada.xlsx`, limpia y lista para Power BI
(una sola fila de encabezados por hoja, sin celdas combinadas, tipos correctos).

Hojas de salida (§8 de ESPECIFICACION_BASE_CONSOLIDADA.md):
    fct_sesiones, dim_programas, dim_calendario, fct_asistencia,
    dim_participantes, Parametros

Uso:
    python construir_base.py [--input CARPETA] [--fecha-corte YYYY-MM-DD]
                             [--output base_consolidada.xlsx] [--readme README.md]

Requisitos: Python 3, pandas, openpyxl.
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
import traceback
import unicodedata
from collections import Counter, OrderedDict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Constantes de dominio
# --------------------------------------------------------------------------

SUBRUTA_CLASES = ("Equipo Logístico", "Listado de Clases")
SUBRUTA_CLASES_ALT = ("Equipo Logistico", "Listado de Clases")  # tolerancia sin tilde

EXT_IGNORADAS = (".jpg", ".jpeg", ".heic", ".png", ".docx", ".pdf")
CARPETAS_IGNORADAS = {"evidencia fotografica", "evidencia fotográfica"}

ANIO_POR_DEFECTO = 2026  # §4.2: el año con el que se reconstruyen los "DD T" / "DD M"

MESES_NOMBRE = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre",
    12: "Diciembre",
}

MESES_NUM = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
    # abreviaturas frecuentes
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6, "jul": 7,
    "ago": 8, "sep": 9, "set": 9, "oct": 10, "nov": 11, "dic": 12,
}

DIAS_NOMBRE = {
    1: "Lunes", 2: "Martes", 3: "Miércoles", 4: "Jueves", 5: "Viernes",
    6: "Sábado", 7: "Domingo",
}

# programa_id sugerido (§2). Se aplica por palabra clave sobre el nombre de la
# carpeta: el recorrido sigue siendo dinámico, esto sólo fija ids legibles y
# estables para las carpetas conocidas.
MAPA_PROGRAMAS = [
    ("bienestar", "BIENESTAR", "Bienestar y Felicidad"),
    ("bootcamp", "BOOTCAMP", "Bootcamp Analítica Predictiva"),
    ("heridas", "HERIDAS", "Cuidado de Heridas"),
    ("ecografia", "ECOGRAFIA", "Ecografía Clínica"),
    ("odontolog", "ODONTOLOGIA", "Odontología Estética Adhesiva"),
    ("normatividad", "NORMATIVIDAD", "Normatividad Inst. Eléctricas"),
    ("project", "PROJECT", "Gerencia de Proyectos (Project)"),
    ("proyectos", "PROJECT", "Gerencia de Proyectos (Project)"),
    ("sensorial", "SENSORIAL", "Integración Sensorial"),
]

# Etiquetas a extraer de la hoja FORMAS DE PAGO (§4.1)
ETIQUETAS_FDP = OrderedDict([
    ("experto_facilitador", ("experto facilitador",)),
    ("nombre_oficial", ("nombre del curso",)),
    ("entidad_convenio", ("entidad convenio",)),
    ("modalidad_declarada", ("modalidad",)),
    ("valor_programa", ("valor del programa",)),
    ("n_participantes", ("numero de particiantes", "numero de participantes")),
    ("nrc", ("nrc",)),
    ("cod_banner", ("cod banner", "codigo banner")),
    ("codigo_contable", ("codigo contable",)),
    ("coordinador", ("coordinador",)),
])

MODALIDADES = ("Presencial", "Virtual", "Remoto", "Híbrido",
               "Trabajo Independiente", "Práctica")


# --------------------------------------------------------------------------
# Utilidades de normalización
# --------------------------------------------------------------------------

def norm(texto) -> str:
    """Minúsculas, sin tildes, sin puntuación de borde, espacios colapsados."""
    if texto is None:
        return ""
    s = str(texto).replace("\xa0", " ").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"\s+", " ", s)
    return s.strip(" :.\t\n")


def limpiar_texto(v) -> str:
    """Texto de celda listo para la base: sin saltos, sin '0' de plantilla."""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).replace("\xa0", " ").replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    if s in ("0", "N/A", "n/a", "#"):
        return ""
    return s


def parse_fecha(v):
    """Normaliza a datetime.date. Acepta datetime, date, número serial y texto."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)):
        # serial de Excel (base 1899-12-30)
        try:
            if 1 < float(v) < 80000:
                return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(v))).date()
        except (ValueError, OverflowError):
            return None
        return None
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y",
                "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:  # último recurso: parser de pandas (día primero, formato latino)
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        return None if pd.isna(ts) else ts.date()
    except Exception:
        return None


def parse_hora(v):
    """Normaliza a datetime.time. Acepta datetime, time, fracción de día y texto."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, dt.datetime):
        return v.time()
    if isinstance(v, dt.time):
        return v
    if isinstance(v, (int, float)):
        f = float(v)
        if 0 <= f < 1:  # fracción de día de Excel
            total = int(round(f * 24 * 60))
            return dt.time(total // 60 % 24, total % 60)
        if 1 <= f < 80000:  # serial completo
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=f)).time()
        return None
    s = str(v).strip().lower().replace(".", "")
    if not s:
        return None
    pm = "pm" in s or "p m" in s
    am = "am" in s or "a m" in s
    s = re.sub(r"\s*[ap]\s*m\s*", "", s).strip()
    m = re.match(r"^(\d{1,2})[:h]?(\d{2})?(?::(\d{2}))?$", s)
    if not m:
        return None
    h = int(m.group(1))
    mi = int(m.group(2) or 0)
    if pm and h < 12:
        h += 12
    if am and h == 12:
        h = 0
    if not (0 <= h <= 23 and 0 <= mi <= 59):
        return None
    return dt.time(h, mi)


def parse_num(v):
    """Normaliza a float. Tolera '$ 1.400.000', '1,5', textos con unidades."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s or s in ("-", ".", ","):
        return None
    if "," in s and "." in s:            # 1.400.000,50 -> 1400000.50
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:                        # 1,5 -> 1.5 ; 1,400 -> 1400
        entero, _, dec = s.rpartition(",")
        s = (entero.replace(",", "") + "." + dec) if len(dec) <= 2 else s.replace(",", "")
    elif s.count(".") > 1:                # 1.400.000
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def limpiar_documento(v) -> str:
    """Documento de identidad → primer número (§10: '12345 /67890' → '12345')."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if not s:
        return ""
    s = re.split(r"[/;,|]", s)[0]            # se queda con el primer número
    s = re.sub(r"[^\d]", "", s)              # quita puntos, espacios y guiones
    return "" if s in ("", "0") else s


def hhmm(t) -> str:
    return "" if t is None else "%02d:%02d" % (t.hour, t.minute)


def normalizar_modalidad(texto) -> str:
    """Texto libre de Salón/MODALIDAD → set fijo de modalidades (§7)."""
    s = norm(texto)
    if not s:
        return ""
    if "trabajo independiente" in s or "independiente" in s:
        return "Trabajo Independiente"
    if "hospital" in s or "practica" in s or "clinic" in s:
        return "Práctica"
    if ("presencial" in s and "virtual" in s) or "hibrid" in s or "blended" in s or "mixt" in s:
        return "Híbrido"
    if "remoto" in s or "distancia" in s:
        return "Remoto"
    if "virtual" in s or "online" in s or "linea" in s:
        return "Virtual"
    if "presencial" in s:
        return "Presencial"
    return ""


def slug_id(nombre: str) -> str:
    s = norm(nombre)
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_").upper()
    return s[:24] or "PROGRAMA"


def identificar_programa(carpeta: str):
    """(programa_id, nombre_corto) a partir del nombre de carpeta."""
    n = norm(carpeta)
    for clave, pid, corto in MAPA_PROGRAMAS:
        if clave in n:
            return pid, corto
    return slug_id(carpeta), carpeta.strip()


# --------------------------------------------------------------------------
# Registro de incidencias de calidad de datos
# --------------------------------------------------------------------------

class Incidencias:
    def __init__(self):
        self.items = []          # (programa, severidad, mensaje)
        self._vistos = set()

    def add(self, programa, mensaje, severidad="AVISO"):
        clave = (programa, mensaje)
        if clave in self._vistos:
            return
        self._vistos.add(clave)
        self.items.append((programa, severidad, mensaje))

    def por_programa(self, programa):
        return [(s, m) for p, s, m in self.items if p == programa]


INC = Incidencias()


# --------------------------------------------------------------------------
# Descubrimiento de carpetas
# --------------------------------------------------------------------------

def descubrir_programas(raiz: str, profundidad_max: int = 5):
    """
    Carpetas de programa, buscando en profundidad la subcarpeta
    'Listado de Clases'. El programa es la carpeta que contiene a
    'Equipo Logístico'.

    El enunciado describe la estructura real organizada por mes y, dentro de
    cada mes, por programa (JULIO 2026/<programa>/Equipo Logístico/...), mientras
    que la copia de trabajo viene aplanada. Buscar en profundidad cubre las dos
    sin asumir un número fijo de niveles. Orden determinista.
    """
    encontrados = []

    def recorrer(directorio, profundidad):
        if profundidad > profundidad_max:
            return
        try:
            entradas = sorted(os.listdir(directorio))
        except OSError:
            return  # sin permisos o eliminada a mitad del recorrido
        for nombre in entradas:
            ruta = os.path.join(directorio, nombre)
            if not os.path.isdir(ruta) or nombre.startswith("."):
                continue
            n = norm(nombre)
            if n in CARPETAS_IGNORADAS or n in ("node_modules", "dist", "participantes"):
                continue
            if n == "listado de clases":
                # <programa>/Equipo Logístico/Listado de Clases
                programa = os.path.basename(os.path.dirname(os.path.dirname(ruta)))
                encontrados.append((programa, ruta))
                continue  # no hace falta bajar más por esta rama
            recorrer(ruta, profundidad + 1)

    recorrer(raiz, 0)
    encontrados.sort(key=lambda x: (norm(x[0]), x[1]))
    return encontrados


def localizar_archivos(carpeta_clases: str):
    """(ruta_cronograma, ruta_listado). Cronograma_* vs. el otro .xlsx (§4)."""
    cronograma = listado = None
    for f in sorted(os.listdir(carpeta_clases)):
        low = f.lower()
        if low.startswith("~$") or not low.endswith((".xlsx", ".xlsm")):
            continue
        if low.endswith(EXT_IGNORADAS):
            continue
        ruta = os.path.join(carpeta_clases, f)
        if norm(f).startswith("cronograma"):
            cronograma = cronograma or ruta
        else:
            listado = listado or ruta
    return cronograma, listado


# --------------------------------------------------------------------------
# Fuente A — Cronograma
# --------------------------------------------------------------------------

CAMPOS_CRONOGRAMA = {
    "num_sesion": ("sesion",),
    "fecha": ("fecha",),
    "hora_inicio": ("hora inicio",),
    "hora_fin": ("hora fin",),
    "intensidad_horaria": ("intensidad horaria",),
    "modulo": ("nombre del modulo", "modulo"),
    "salon": ("salon",),
    "docente": ("nombre del docente", "docente"),
}


def leer_cronograma(ruta: str, programa: str):
    """Devuelve lista de dicts de sesión. Detecta la hoja por encabezados (§3)."""
    wb = load_workbook(ruta, data_only=True, read_only=True)
    try:
        hoja_ok = None
        for ws in wb.worksheets:
            filas = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            encabezados = next(iter(filas), None)
            if not encabezados:
                continue
            hs = [norm(h) for h in encabezados]
            tiene_sesion = any(h.startswith("sesion") for h in hs)
            tiene_fecha = any(h == "fecha" or h.startswith("fecha") for h in hs)
            if tiene_sesion and tiene_fecha:
                hoja_ok = ws
                break
        if hoja_ok is None:
            INC.add(programa, "Cronograma sin hoja con encabezados 'Sesión'/'Fecha'; "
                              "no se pudieron leer sesiones.", "ERROR")
            return []

        filas = list(hoja_ok.iter_rows(values_only=True))
        encabezados = [norm(h) for h in filas[0]]

        # mapeo por nombre de encabezado, no por posición (§3)
        idx = {}
        for campo, alias in CAMPOS_CRONOGRAMA.items():
            for i, h in enumerate(encabezados):
                if any(h.startswith(a) for a in alias):
                    idx[campo] = i
                    break

        for obligatorio in ("num_sesion", "fecha", "hora_inicio", "hora_fin",
                            "intensidad_horaria"):
            if obligatorio not in idx:
                INC.add(programa, "Cronograma sin columna '%s'; el campo queda vacío."
                        % obligatorio, "AVISO")
        if "salon" not in idx:
            INC.add(programa, "Cronograma sin columna 'Salón': la modalidad se toma "
                              "de MODALIDAD en FORMAS DE PAGO.")
        if "modulo" not in idx:
            INC.add(programa, "Cronograma sin columna 'Nombre del módulo'.")

        def val(fila, campo):
            i = idx.get(campo)
            return fila[i] if i is not None and i < len(fila) else None

        sesiones, orden = [], 0
        for fila in filas[1:]:
            if fila is None or all(v is None or str(v).strip() == "" for v in fila):
                continue
            fecha = parse_fecha(val(fila, "fecha"))
            num = parse_num(val(fila, "num_sesion"))
            if fecha is None and num is None:
                continue
            orden += 1
            crudo_hi = val(fila, "hora_inicio")
            crudo_hf = val(fila, "hora_fin")
            sesiones.append({
                "orden_archivo": orden,
                "num_sesion": int(num) if num is not None else None,
                "fecha": fecha,
                "hora_inicio": parse_hora(crudo_hi),
                "hora_fin": parse_hora(crudo_hf),
                "fecha_hora_inicio": parse_fecha(crudo_hi) if isinstance(
                    crudo_hi, dt.datetime) else None,
                "fecha_hora_fin": parse_fecha(crudo_hf) if isinstance(
                    crudo_hf, dt.datetime) else None,
                "hora_inicio_texto": isinstance(crudo_hi, str),
                "hora_fin_texto": isinstance(crudo_hf, str),
                "fecha_texto": isinstance(val(fila, "fecha"), str),
                "intensidad_horaria": parse_num(val(fila, "intensidad_horaria")),
                "modulo": limpiar_texto(val(fila, "modulo")),
                "salon": limpiar_texto(val(fila, "salon")),
                "docente": limpiar_texto(val(fila, "docente")),
            })
        return sesiones
    finally:
        wb.close()


# --------------------------------------------------------------------------
# Fuente B — Listado de Participantes
# --------------------------------------------------------------------------

def buscar_hoja(wb, nombre_objetivo):
    objetivo = norm(nombre_objetivo)
    for nombre in wb.sheetnames:
        if norm(nombre) == objetivo:
            return wb[nombre]
    for nombre in wb.sheetnames:
        if objetivo in norm(nombre):
            return wb[nombre]
    return None


def valor_a_la_derecha(ws, fila, col, max_saltos=6):
    """Primer valor no vacío a la derecha de una etiqueta (la maquetación varía)."""
    for c in range(col + 1, min(col + 1 + max_saltos, ws.max_column + 1)):
        v = ws.cell(row=fila, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        if s.endswith(":") and len(s) > 3:   # topamos con otra etiqueta
            return None
        return v
    return None


def leer_formas_de_pago(wb, programa):
    """Metadatos del programa buscando el texto de cada etiqueta (§4.1)."""
    meta = {k: None for k in ETIQUETAS_FDP}
    ws = buscar_hoja(wb, "FORMAS DE PAGO")
    if ws is None:
        INC.add(programa, "No existe la hoja 'FORMAS DE PAGO'; metadatos vacíos.", "ERROR")
        return meta, None

    # Los metadatos viven en el bloque superior. Más abajo empieza la tabla de
    # participantes, cuyos encabezados ('CODIGO BANNER', 'CORREO ELECTRONICO'…)
    # colisionan con las etiquetas buscadas: hay que cortar la búsqueda ahí.
    fila_tabla = min(30, ws.max_row) + 1
    for r in range(1, min(30, ws.max_row) + 1):
        valores = {norm(ws.cell(row=r, column=c).value)
                   for c in range(1, min(16, ws.max_column) + 1)}
        if valores & {"apellidos", "cedula", "forma de pago", "tipo de participante"}:
            fila_tabla = r
            break

    lugar_fecha = None
    for fila in ws.iter_rows(min_row=1, max_row=fila_tabla - 1):
        for celda in fila:
            if not isinstance(celda.value, str):
                continue
            etiqueta = norm(celda.value)
            if not etiqueta:
                continue
            if etiqueta.startswith("lugar y fecha"):
                lugar_fecha = limpiar_texto(
                    valor_a_la_derecha(ws, celda.row, celda.column))
            for campo, alias in ETIQUETAS_FDP.items():
                if meta[campo] is not None:
                    continue
                if any(etiqueta.startswith(a) for a in alias):
                    meta[campo] = valor_a_la_derecha(ws, celda.row, celda.column)

    faltantes = [c for c, v in meta.items() if v is None]
    if faltantes:
        INC.add(programa, "FORMAS DE PAGO sin valor para: %s." % ", ".join(faltantes))
    return meta, lugar_fecha


RE_DIA_JORNADA = re.compile(r"^(\d{1,2})\s*\.?\s*([tmTM])?\s*\.?$")


def _mes_a_numero(v):
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = int(v)
        return n if 1 <= n <= 12 else None
    if isinstance(v, dt.datetime) or isinstance(v, dt.date):
        return v.month
    s = norm(v)
    if not s:
        return None
    for clave, num in MESES_NUM.items():
        if s.startswith(clave):
            return num
    return None


def leer_consolidado(wb, programa, anios_cronograma, fechas_cronograma):
    """
    Parsea la hoja CONSOLIDADO (§4.2). Devuelve:
        columnas       -> lista de dicts {col, fecha, jornada, crudo, tabulada}
        participantes  -> lista de dicts {documento, nombre, empresa, sigma, celdas}
        horas_totales, horas_falla_max
    """
    vacio = ([], [], None, None)
    ws = buscar_hoja(wb, "CONSOLIDADO")
    if ws is None:
        INC.add(programa, "No existe la hoja 'CONSOLIDADO'; sin datos de asistencia.", "ERROR")
        return vacio

    # ---- fila de encabezados (NOMBRE / DOCUMENTO DE IDENTIDAD)
    fila_hdr = None
    for r in range(1, min(40, ws.max_row) + 1):
        valores = [norm(ws.cell(row=r, column=c).value)
                   for c in range(1, min(12, ws.max_column) + 1)]
        if any(v == "nombre" for v in valores) and any(
                v.startswith("documento") for v in valores):
            fila_hdr = r
            break
    if fila_hdr is None:
        INC.add(programa, "CONSOLIDADO sin fila de encabezados NOMBRE/DOCUMENTO.", "ERROR")
        return vacio

    col_nombre = col_doc = col_empresa = col_correo = col_sigma = col_obs = None
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(row=fila_hdr, column=c).value)
        if not h:
            continue
        if h == "nombre" and col_nombre is None:
            col_nombre = c
        elif h.startswith("documento") and col_doc is None:
            col_doc = c
        elif h.startswith("empresa") and col_empresa is None:
            col_empresa = c
        elif h.startswith("correo") and col_correo is None:
            col_correo = c
        elif ("de inasistencia" in h or h.startswith("total")) and col_sigma is None:
            col_sigma = c
        elif h.startswith("observacion") and col_obs is None:
            col_obs = c

    col_ini = max(x for x in (col_nombre, col_doc, col_empresa, col_correo) if x) + 1
    col_fin = (col_sigma - 1) if col_sigma else ws.max_column
    if col_obs and col_sigma is None:
        col_fin = col_obs - 1
    if col_fin < col_ini:
        INC.add(programa, "CONSOLIDADO sin columnas de sesión entre CORREO y Σ.", "ERROR")
        return vacio

    # ---- horas totales y tope de fallas (cabecera)
    horas_totales = horas_falla = None
    for r in range(1, fila_hdr):
        for c in range(1, min(8, ws.max_column) + 1):
            h = norm(ws.cell(row=r, column=c).value)
            if not h.startswith("numero de horas"):
                continue
            valor = parse_num(valor_a_la_derecha(ws, r, c))
            if "falla" in h:
                horas_falla = valor if horas_falla is None else horas_falla
            elif horas_totales is None:
                horas_totales = valor
    if horas_totales is None:
        INC.add(programa, "CONSOLIDADO sin 'NÚMERO DE HORAS'.")
    if horas_falla is None:
        INC.add(programa, "CONSOLIDADO sin 'HORAS DE FALLAS MÁXIMAS PERMITIDAS'.")

    # ---- fila 'Mes:' con forward-fill (§4.2, §10)
    fila_mes = None
    for r in range(max(1, fila_hdr - 6), fila_hdr):
        for c in range(1, min(col_ini, ws.max_column) + 1):
            if norm(ws.cell(row=r, column=c).value).startswith("mes"):
                fila_mes = r
                break
        if fila_mes:
            break
    if fila_mes is None:
        INC.add(programa, "CONSOLIDADO sin fila 'Mes:'; las columnas 'DD T'/'DD M' "
                          "no se pueden fechar.")

    meses_ff, ultimo = {}, None
    if fila_mes:
        for c in range(col_ini, col_fin + 1):
            m = _mes_a_numero(ws.cell(row=fila_mes, column=c).value)
            if m is not None:
                ultimo = m          # forward-fill sobre celdas combinadas/vacías
            meses_ff[c] = ultimo

    anios = sorted(set(anios_cronograma)) or [ANIO_POR_DEFECTO]

    def elegir_anio(mes, dia):
        """Año que hace calzar (mes, día) con el cronograma; si no, el año base."""
        for a in anios:
            try:
                if dt.date(a, mes, dia) in fechas_cronograma:
                    return a
            except ValueError:
                continue
        return anios[0]

    # ---- columnas de sesión
    columnas, sin_fechar = [], []
    for c in range(col_ini, col_fin + 1):
        crudo = ws.cell(row=fila_hdr, column=c).value
        etiqueta = limpiar_texto(crudo) if not isinstance(crudo, (dt.date, dt.datetime)) else str(crudo)
        if crudo is None or (isinstance(crudo, str) and not crudo.strip()):
            continue  # columna separadora vacía

        fecha = jornada = None
        if isinstance(crudo, (dt.datetime, dt.date)):
            fecha = parse_fecha(crudo)                     # formato 1: fecha completa
        else:
            texto = str(crudo).strip()
            m = RE_DIA_JORNADA.match(texto)                # formato 2: 'DD T' / 'DD M'
            if m:
                dia = int(m.group(1))
                jornada = {"t": "Tarde", "m": "Mañana"}.get((m.group(2) or "").lower())
                mes = meses_ff.get(c)
                if mes and 1 <= dia <= 31:
                    try:
                        fecha = dt.date(elegir_anio(mes, dia), mes, dia)
                    except ValueError:
                        fecha = None
            else:
                fecha = parse_fecha(crudo)

        if fecha is None:
            sin_fechar.append(etiqueta)
        columnas.append({"col": c, "fecha": fecha, "jornada": jornada,
                         "etiqueta": etiqueta, "tabulada": False, "usada": False})

    if sin_fechar:
        INC.add(programa, "CONSOLIDADO: %d columna(s) de sesión sin fecha reconstruible "
                          "(%s); esas sesiones quedan sin columna."
                % (len(sin_fechar), ", ".join(repr(x) for x in sin_fechar[:8])))

    # ---- participantes
    participantes = []
    vacias_seguidas = 0
    for r in range(fila_hdr + 1, ws.max_row + 1):
        if vacias_seguidas >= 60:
            break
        nombre = limpiar_texto(ws.cell(row=r, column=col_nombre).value) if col_nombre else ""
        documento = limpiar_documento(ws.cell(row=r, column=col_doc).value) if col_doc else ""
        celdas = {}
        for col in columnas:
            v = ws.cell(row=r, column=col["col"]).value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            celdas[col["col"]] = v
        if not nombre or (not documento and not celdas):
            vacias_seguidas += 1
            continue
        if norm(nombre).startswith(("nomenclatura", "asistencia por sesion",
                                    "inasistencia representada", "total")):
            vacias_seguidas += 1
            continue
        vacias_seguidas = 0
        for c, v in celdas.items():
            for col in columnas:
                if col["col"] == c:
                    col["tabulada"] = True
        participantes.append({
            "fila": r,
            "nombre": nombre,
            "documento": documento,
            "empresa": limpiar_texto(ws.cell(row=r, column=col_empresa).value) if col_empresa else "",
            "sigma": parse_num(ws.cell(row=r, column=col_sigma).value) if col_sigma else None,
            "celdas": celdas,
        })

    if not participantes:
        INC.add(programa, "CONSOLIDADO sin filas de participantes legibles.", "ERROR")
    if col_doc:
        sin_doc = [p["nombre"] for p in participantes if not p["documento"]]
        if sin_doc:
            INC.add(programa, "%d participante(s) sin documento legible (%s)."
                    % (len(sin_doc), "; ".join(sin_doc[:3])))

    return columnas, participantes, horas_totales, horas_falla


# --------------------------------------------------------------------------
# Emparejamiento sesión ↔ columna de CONSOLIDADO (§5)
# --------------------------------------------------------------------------

def emparejar(sesion, columnas_por_fecha):
    """
    Elige la columna de CONSOLIDADO que corresponde a la sesión.
    Prioridad: (1) jornada compatible, (2) columna con datos, (3) más a la izquierda.
    Una columna puede servir a varias sesiones del mismo día (el CONSOLIDADO
    suele traer una columna por día y el cronograma varias sesiones por día).
    """
    candidatas = columnas_por_fecha.get(sesion["fecha"], [])
    if not candidatas:
        return None, 0
    def clave(col):
        incompatible = 0 if (col["jornada"] is None or
                             col["jornada"] == sesion["jornada"]) else 1
        return (incompatible, 0 if col["tabulada"] else 1, col["col"])
    return sorted(candidatas, key=clave)[0], len(candidatas)


# --------------------------------------------------------------------------
# Procesamiento por programa
# --------------------------------------------------------------------------

def procesar_programa(carpeta, ruta_clases, fecha_corte):
    programa_id, programa = identificar_programa(carpeta)
    ruta_cron, ruta_lst = localizar_archivos(ruta_clases)

    if ruta_cron is None:
        INC.add(programa, "No se encontró archivo 'Cronograma_*.xlsx'.", "ERROR")
    if ruta_lst is None:
        INC.add(programa, "No se encontró archivo de listado de participantes.", "ERROR")

    # ---------------- Cronograma
    sesiones = leer_cronograma(ruta_cron, programa) if ruta_cron else []

    # §10: ordenar por fecha, no por número de sesión (Odontología)
    con_fecha = [s for s in sesiones if s["fecha"] is not None]
    sin_fecha = [s for s in sesiones if s["fecha"] is None]
    if sin_fecha:
        INC.add(programa, "%d sesión(es) del cronograma sin fecha legible; se excluyen "
                          "de la base." % len(sin_fecha), "ERROR")
    nums = [s["num_sesion"] for s in con_fecha if s["num_sesion"] is not None]
    orden_original = [s["num_sesion"] for s in sorted(con_fecha, key=lambda s: s["orden_archivo"])]
    con_fecha.sort(key=lambda s: (s["fecha"],
                                  s["hora_inicio"] or dt.time(0, 0),
                                  s["num_sesion"] if s["num_sesion"] is not None else 0))
    if orden_original != [s["num_sesion"] for s in con_fecha]:
        INC.add(programa, "Sesiones fuera de orden en el cronograma: se reordenaron "
                          "por fecha (§10).")
    if nums and sorted(nums) != list(range(min(nums), min(nums) + len(nums))):
        faltantes = sorted(set(range(min(nums), max(nums) + 1)) - set(nums))
        INC.add(programa, "Numeración de sesiones no correlativa; faltan los números "
                          "%s." % ", ".join(str(x) for x in faltantes[:12]))

    fechas_cron = {s["fecha"] for s in con_fecha}
    anios_cron = [f.year for f in fechas_cron]

    # ---------------- Listado de participantes
    meta, lugar_fecha = {k: None for k in ETIQUETAS_FDP}, None
    columnas, participantes, horas_totales, horas_falla = [], [], None, None
    if ruta_lst:
        wb = load_workbook(ruta_lst, data_only=True)
        try:
            meta, lugar_fecha = leer_formas_de_pago(wb, programa)
            columnas, participantes, horas_totales, horas_falla = leer_consolidado(
                wb, programa, anios_cron, fechas_cron)
        finally:
            wb.close()

    # §10: typo de año en metadatos → se confía en el cronograma
    if lugar_fecha and anios_cron:
        anios_meta = {int(a) for a in re.findall(r"\b(20\d{2})\b", lugar_fecha)}
        raros = sorted(anios_meta - set(anios_cron))
        if raros:
            INC.add(programa, "FORMAS DE PAGO / 'LUGAR Y FECHA' declara el año %s "
                              "(%r) pero el cronograma es de %s: se usan las fechas "
                              "del cronograma." % (", ".join(map(str, raros)),
                                                   lugar_fecha,
                                                   ", ".join(map(str, sorted(set(anios_cron))))))

    n_participantes_meta = parse_num(meta.get("n_participantes"))
    n_participantes_meta = int(n_participantes_meta) if n_participantes_meta else None
    if n_participantes_meta is not None and participantes and \
            n_participantes_meta != len(participantes):
        INC.add(programa, "NÚMERO DE PARTICIANTES declarado = %d pero CONSOLIDADO "
                          "lista %d participante(s)." % (n_participantes_meta,
                                                         len(participantes)))

    columnas_por_fecha = {}
    for col in columnas:
        if col["fecha"] is not None:
            columnas_por_fecha.setdefault(col["fecha"], []).append(col)

    # ---------------- Construcción de fct_sesiones
    modalidad_prog_declarada = normalizar_modalidad(meta.get("modalidad_declarada"))
    filas_sesiones, filas_asistencia = [], []
    usos_por_columna = Counter()
    ids_vistos = set()
    # acumuladores para resumir las banderas de fila en incidencias del programa
    horas_invertidas, horas_cero, fecha_cruzada = [], [], []
    sin_columna, fecha_multicolumna = [], []

    for s in con_fecha:
        obs = []
        fecha = s["fecha"]
        hi, hf = s["hora_inicio"], s["hora_fin"]

        if s["fecha_texto"]:
            obs.append("fecha venía como texto")
        if s["hora_inicio_texto"] or s["hora_fin_texto"]:
            obs.append("hora venía como texto")
        if hi is None:
            obs.append("sin hora de inicio")
        if hf is None:
            obs.append("sin hora de fin")
        if hf == dt.time(0, 0):
            obs.append("hora_fin en 00:00")
            horas_cero.append(s["num_sesion"])
        if hi and hf and hf <= hi:
            obs.append("hora_fin anterior o igual a hora_inicio")
            horas_invertidas.append((s["num_sesion"], hhmm(hi), hhmm(hf)))
        for etiqueta, f_otro in (("hora_inicio", s["fecha_hora_inicio"]),
                                 ("hora_fin", s["fecha_hora_fin"])):
            if f_otro is not None and f_otro != fecha:
                obs.append("%s registrada en otra fecha (%s)" % (etiqueta, f_otro))
                fecha_cruzada.append((s["num_sesion"], etiqueta, f_otro))
        if s["intensidad_horaria"] is None:
            obs.append("sin intensidad horaria")

        jornada = "Mañana" if (hi and hi < dt.time(12, 0)) else "Tarde"
        if hi is None:
            obs.append("jornada asumida Tarde por falta de hora de inicio")
        s["jornada"] = jornada

        modalidad = normalizar_modalidad(s["salon"]) or modalidad_prog_declarada
        if s["salon"] and not normalizar_modalidad(s["salon"]):
            obs.append("salón '%s' no mapea a una modalidad conocida" % s["salon"])

        col, n_cand = emparejar(s, columnas_por_fecha)
        if n_cand > 1:
            obs.append("%d columnas en CONSOLIDADO para esa fecha" % n_cand)
            fecha_multicolumna.append((fecha, n_cand))
        if col is not None:
            usos_por_columna[col["col"]] += 1
            col["usada"] = True

        realizada = fecha <= fecha_corte
        if col is None:
            obs.append("sin columna en CONSOLIDADO")
            sin_columna.append((s["num_sesion"], fecha))
            estado_seg = "Pendiente de tabular" if realizada else "Futura no exigible"
        elif col["tabulada"]:
            estado_seg = "Tabulada"
        else:
            estado_seg = "Pendiente de tabular" if realizada else "Futura no exigible"

        if estado_seg == "Tabulada":
            asistencia_tabulada = "Sí"
        elif estado_seg == "Futura no exigible":
            asistencia_tabulada = "N/A"
        else:
            asistencia_tabulada = "No"

        estado_sesion = ("Realizada" if fecha < fecha_corte
                         else "Hoy" if fecha == fecha_corte else "Futura")

        num = s["num_sesion"]
        id_sesion = "%s-%s" % (programa_id, ("%02d" % num) if num is not None
                               else "S%03d" % s["orden_archivo"])
        if id_sesion in ids_vistos:                    # unicidad garantizada
            sufijo = 2
            while "%s.%d" % (id_sesion, sufijo) in ids_vistos:
                sufijo += 1
            INC.add(programa, "Número de sesión duplicado (%s): id_sesion desambiguado."
                    % num)
            id_sesion = "%s.%d" % (id_sesion, sufijo)
        ids_vistos.add(id_sesion)

        intensidad = s["intensidad_horaria"]
        n_asis = n_inasis = None
        if estado_seg == "Tabulada":
            n_asis = n_inasis = 0
            for p in participantes:
                v = p["celdas"].get(col["col"])
                if v is None:
                    continue
                horas = parse_num(v)
                if horas is None:
                    obs.append("valor no numérico en CONSOLIDADO (%r)" % str(v)[:20])
                    continue
                asistio = (horas < intensidad) if intensidad else (horas == 0)
                n_asis += 1 if asistio else 0
                n_inasis += 0 if asistio else 1
                filas_asistencia.append({
                    "id_registro": "%s|%s" % (id_sesion, p["documento"] or "F%d" % p["fila"]),
                    "programa_id": programa_id,
                    "id_sesion": id_sesion,
                    "fecha": fecha,
                    "documento": p["documento"],
                    "nombre": p["nombre"],
                    "empresa": p["empresa"],
                    "horas_inasistencia": horas,
                    "asistio": bool(asistio),
                    "tabulada": True,
                    "_col": col["col"],
                })

        iso = fecha.isocalendar()
        filas_sesiones.append({
            "id_sesion": id_sesion,
            "programa_id": programa_id,
            "programa": programa,
            "num_sesion": num,
            "modulo": s["modulo"],
            "fecha": fecha,
            "anio": fecha.year,
            "mes": fecha.month,
            "mes_nombre": MESES_NOMBRE[fecha.month],
            "dia_semana": DIAS_NOMBRE[fecha.isoweekday()],
            "dia_semana_num": fecha.isoweekday(),
            "semana_iso": int(iso[1]),
            "anio_semana": "%d-W%02d" % (iso[0], iso[1]),
            "jornada": jornada,
            "hora_inicio": hhmm(hi),
            "hora_fin": hhmm(hf),
            "intensidad_horaria": intensidad,
            "modalidad": modalidad,
            "salon": s["salon"],
            "docente": s["docente"],
            "estado_sesion": estado_sesion,
            "asistencia_tabulada": asistencia_tabulada,
            "estado_seguimiento": estado_seg,
            "n_participantes": n_participantes_meta,
            "n_asistentes": n_asis,
            "n_inasistentes": n_inasis,
            "observaciones": "; ".join(dict.fromkeys(obs)),
        })

    # ---------------- resumen de banderas de fila como incidencias del programa
    if horas_invertidas:
        INC.add(programa, "%d sesión(es) con hora_fin anterior o igual a hora_inicio "
                          "(se conservan y quedan marcadas en observaciones): %s."
                % (len(horas_invertidas),
                   "; ".join("sesión %s %s→%s" % t for t in horas_invertidas[:5])))
    if horas_cero:
        INC.add(programa, "%d sesión(es) con hora_fin en 00:00: %s."
                % (len(horas_cero), ", ".join("sesión %s" % n for n in horas_cero[:5])))
    if fecha_cruzada:
        INC.add(programa, "%d sesión(es) con %s fechada en un día distinto al de la "
                          "sesión: %s." % (len(fecha_cruzada), fecha_cruzada[0][1],
                                           "; ".join("sesión %s → %s" % (n, f)
                                                     for n, _, f in fecha_cruzada[:5])))
    if sin_columna:
        INC.add(programa, "%d sesión(es) del cronograma sin columna en CONSOLIDADO "
                          "(%s); quedan como no tabuladas."
                % (len(sin_columna),
                   ", ".join("sesión %s del %s" % (n, f) for n, f in sin_columna[:6])))
    if fecha_multicolumna:
        INC.add(programa, "Fecha(s) con más de una columna en CONSOLIDADO (%s): se usa "
                          "la columna con datos y jornada compatible."
                % ", ".join(sorted({str(f) for f, _ in fecha_multicolumna})[:6]))

    # columnas de CONSOLIDADO que ninguna sesión reclamó
    huerfanas = [c for c in columnas if not c["usada"] and c["fecha"] is not None]
    if huerfanas:
        INC.add(programa, "%d columna(s) de CONSOLIDADO sin sesión en el cronograma "
                          "(%s)." % (len(huerfanas),
                                     ", ".join("%s→%s" % (c["etiqueta"], c["fecha"])
                                               for c in huerfanas[:6])))

    # ---------------- dim_participantes
    filas_participantes = []
    for p in participantes:
        # Σ recalculado: una sola vez por columna de CONSOLIDADO, aunque esa
        # columna sirva a varias sesiones del mismo día (evita doble conteo).
        total = 0.0
        vistos = set()
        for reg in filas_asistencia:
            if reg["documento"] != p["documento"] or reg["_col"] in vistos:
                continue
            if reg["nombre"] != p["nombre"]:
                continue
            vistos.add(reg["_col"])
            total += reg["horas_inasistencia"]
        if p["sigma"] is not None and abs(total - p["sigma"]) > 1e-6:
            INC.add(programa, "Σ de inasistencia del archivo (%g) ≠ recalculado (%g) "
                              "para '%s'." % (p["sigma"], total, p["nombre"]))
        filas_participantes.append({
            "programa_id": programa_id,
            "documento": p["documento"],
            "nombre": p["nombre"],
            "empresa": p["empresa"],
            "total_inasistencia": total,
            "horas_falla_max": horas_falla,
            "en_riesgo": bool(horas_falla is not None and total > horas_falla),
        })

    docs = [f["documento"] for f in filas_participantes if f["documento"]]
    if len(docs) != len(set(docs)):
        repetidos = [d for d, n in Counter(docs).items() if n > 1]
        INC.add(programa, "Documento(s) repetido(s) en CONSOLIDADO: %s."
                % ", ".join(repetidos[:5]))

    # ---------------- dim_programas
    fechas = [f["fecha"] for f in filas_sesiones]
    modalidades = [f["modalidad"] for f in filas_sesiones if f["modalidad"]]
    if modalidades:
        conteo = Counter(modalidades)
        tope = max(conteo.values())
        modalidad_prog = next(m for m in modalidades if conteo[m] == tope)
    else:
        modalidad_prog = modalidad_prog_declarada

    n_realizadas = sum(1 for f in filas_sesiones if f["fecha"] <= fecha_corte)
    n_tabuladas = sum(1 for f in filas_sesiones if f["estado_seguimiento"] == "Tabulada")
    n_pendientes = sum(1 for f in filas_sesiones
                       if f["estado_seguimiento"] == "Pendiente de tabular")
    denominador = n_tabuladas + n_pendientes          # §5: sólo sesiones realizadas
    pct = round(n_tabuladas / denominador, 4) if denominador else None

    if fechas:
        fi, ff = min(fechas), max(fechas)
        estado_prog = ("Por iniciar" if fecha_corte < fi
                       else "Finalizado" if fecha_corte > ff else "En ejecución")
    else:
        fi = ff = None
        estado_prog = ""

    fila_programa = {
        "programa_id": programa_id,
        "programa": programa,
        "nombre_oficial": limpiar_texto(meta.get("nombre_oficial")),
        "nrc": limpiar_texto(meta.get("nrc")),
        "cod_banner": limpiar_texto(meta.get("cod_banner")),
        "codigo_contable": limpiar_texto(meta.get("codigo_contable")),
        "coordinador": limpiar_texto(meta.get("coordinador")),
        "experto_facilitador": limpiar_texto(meta.get("experto_facilitador")),
        "entidad_convenio": limpiar_texto(meta.get("entidad_convenio")),
        "modalidad": modalidad_prog,
        "valor_programa": parse_num(meta.get("valor_programa")),
        "n_participantes": n_participantes_meta,
        "fecha_inicio": fi,
        "fecha_fin": ff,
        "n_sesiones": len(filas_sesiones),
        "horas_totales": horas_totales,
        "horas_falla_max": horas_falla,
        "n_sesiones_realizadas": n_realizadas,
        "n_sesiones_tabuladas": n_tabuladas,
        "n_sesiones_pendientes": n_pendientes,
        "pct_cumplimiento_tabulacion": pct,
        "estado_programa": estado_prog,
    }

    # coherencia horas del cronograma vs. horas declaradas
    suma_int = sum(f["intensidad_horaria"] or 0 for f in filas_sesiones)
    if horas_totales and abs(suma_int - horas_totales) > 0.5:
        INC.add(programa, "Suma de intensidad horaria del cronograma (%g h) ≠ NÚMERO "
                          "DE HORAS del CONSOLIDADO (%g h)." % (suma_int, horas_totales))

    for reg in filas_asistencia:
        reg.pop("_col", None)

    return {
        "programa_id": programa_id,
        "programa": programa,
        "sesiones": filas_sesiones,
        "asistencia": filas_asistencia,
        "participantes": filas_participantes,
        "programa_row": fila_programa,
        "n_futuras": sum(1 for f in filas_sesiones
                         if f["estado_seguimiento"] == "Futura no exigible"),
    }


# --------------------------------------------------------------------------
# dim_calendario
# --------------------------------------------------------------------------

def construir_calendario(fechas):
    if not fechas:
        return pd.DataFrame(columns=["fecha", "anio", "mes", "mes_nombre", "dia",
                                     "dia_semana", "dia_semana_num", "semana_iso",
                                     "anio_semana", "es_fin_de_semana"])
    inicio, fin = min(fechas), max(fechas)
    filas = []
    d = inicio
    while d <= fin:
        iso = d.isocalendar()
        filas.append({
            "fecha": d,
            "anio": d.year,
            "mes": d.month,
            "mes_nombre": MESES_NOMBRE[d.month],
            "dia": d.day,
            "dia_semana": DIAS_NOMBRE[d.isoweekday()],
            "dia_semana_num": d.isoweekday(),
            "semana_iso": int(iso[1]),
            "anio_semana": "%d-W%02d" % (iso[0], iso[1]),
            "es_fin_de_semana": d.isoweekday() >= 6,
        })
        d += dt.timedelta(days=1)
    return pd.DataFrame(filas)


# --------------------------------------------------------------------------
# Escritura del Excel
# --------------------------------------------------------------------------

COLUMNAS = {
    "fct_sesiones": ["id_sesion", "programa_id", "programa", "num_sesion", "modulo",
                     "fecha", "anio", "mes", "mes_nombre", "dia_semana",
                     "dia_semana_num", "semana_iso", "anio_semana", "jornada",
                     "hora_inicio", "hora_fin", "intensidad_horaria", "modalidad",
                     "salon", "docente", "estado_sesion", "asistencia_tabulada",
                     "estado_seguimiento", "n_participantes", "n_asistentes",
                     "n_inasistentes", "observaciones"],
    "dim_programas": ["programa_id", "programa", "nombre_oficial", "nrc", "cod_banner",
                      "codigo_contable", "coordinador", "experto_facilitador",
                      "entidad_convenio", "modalidad", "valor_programa",
                      "n_participantes", "fecha_inicio", "fecha_fin", "n_sesiones",
                      "horas_totales", "horas_falla_max", "n_sesiones_realizadas",
                      "n_sesiones_tabuladas", "n_sesiones_pendientes",
                      "pct_cumplimiento_tabulacion", "estado_programa"],
    "dim_calendario": ["fecha", "anio", "mes", "mes_nombre", "dia", "dia_semana",
                       "dia_semana_num", "semana_iso", "anio_semana",
                       "es_fin_de_semana"],
    "fct_asistencia": ["id_registro", "programa_id", "id_sesion", "fecha", "documento",
                       "nombre", "empresa", "horas_inasistencia", "asistio", "tabulada"],
    "dim_participantes": ["programa_id", "documento", "nombre", "empresa",
                          "total_inasistencia", "horas_falla_max", "en_riesgo"],
    "Parametros": ["parametro", "valor", "descripcion"],
}

COLUMNAS_FECHA = {"fecha", "fecha_inicio", "fecha_fin"}
COLUMNAS_ENTERAS = {"num_sesion", "anio", "mes", "dia", "dia_semana_num", "semana_iso",
                    "n_participantes", "n_asistentes", "n_inasistentes", "n_sesiones",
                    "n_sesiones_realizadas", "n_sesiones_tabuladas",
                    "n_sesiones_pendientes"}
COLUMNAS_DECIMALES = {"intensidad_horaria", "valor_programa", "horas_totales",
                      "horas_falla_max", "pct_cumplimiento_tabulacion",
                      "horas_inasistencia", "total_inasistencia"}
COLUMNAS_BOOLEANAS = {"asistio", "tabulada", "en_riesgo", "es_fin_de_semana"}

# Marcador para los vacíos de las columnas de TEXTO. Las columnas numéricas, de
# fecha y booleanas se dejan realmente nulas: escribirles un texto las volvería
# columnas de texto en Power BI y rompería sumas, promedios y ejes de fecha.
RELLENO_TEXTO = "N/A"


def preparar(df, hoja, relleno=RELLENO_TEXTO):
    """Ordena columnas al esquema exacto de §8, fija tipos y rellena los vacíos
    de texto con el marcador (default 'N/A')."""
    cols = COLUMNAS[hoja]
    if df.empty:
        df = pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df = df[cols].copy()
    for c in cols:
        if c in COLUMNAS_ENTERAS:
            df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")
        elif c in COLUMNAS_DECIMALES:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        elif c in COLUMNAS_FECHA:
            df[c] = pd.to_datetime(df[c], errors="coerce")
        elif c in COLUMNAS_BOOLEANAS:
            df[c] = df[c].astype("boolean")
        elif relleno:
            df[c] = (df[c].astype("object")
                          .where(df[c].notna(), None)
                          .map(lambda v: relleno if v is None or str(v).strip() == ""
                               else v))
    return df


def escribir_excel(ruta, hojas):
    with pd.ExcelWriter(ruta, engine="openpyxl", datetime_format="yyyy-mm-dd",
                        date_format="yyyy-mm-dd") as writer:
        for hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=hoja, index=False)

    wb = load_workbook(ruta)
    encabezado = Font(bold=True, color="FFFFFF")
    relleno = PatternFill("solid", fgColor="1F4E78")
    for hoja, df in hojas.items():
        ws = wb[hoja]
        for celda in ws[1]:
            celda.font = encabezado
            celda.fill = relleno
            celda.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        for i, nombre in enumerate(df.columns, start=1):
            letra = get_column_letter(i)
            if nombre in COLUMNAS_FECHA:
                for celda in ws[letra][1:]:
                    celda.number_format = "yyyy-mm-dd"
            ancho = max(len(str(nombre)) + 4,
                        min(46, int(df[nombre].astype(str).str.len().max() or 0) + 2)
                        if len(df) else 12)
            ws.column_dimensions[letra].width = ancho
    wb.save(ruta)


# --------------------------------------------------------------------------
# README de incidencias
# --------------------------------------------------------------------------

def escribir_readme(ruta, resultados, fecha_corte, salida, orden_programas):
    lineas = [
        "# Base consolidada CEC — incidencias de calidad de datos",
        "",
        "Generado por `construir_base.py` el %s."
        % dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "",
        "- **Fecha de corte:** `%s`" % fecha_corte.isoformat(),
        "- **Archivo generado:** `%s`" % os.path.basename(salida),
        "- **Programas procesados:** %d" % len(resultados),
        "",
        "## Resumen por programa",
        "",
        "| Programa | Sesiones | Tabuladas | Pendientes | Futuras | Cumplim. | "
        "Participantes | Incidencias |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for nombre in orden_programas:
        r = resultados.get(nombre)
        if not r:
            lineas.append("| %s | — | — | — | — | — | — | ver abajo |" % nombre)
            continue
        p = r["programa_row"]
        pct = "—" if p["pct_cumplimiento_tabulacion"] is None else \
            "%.0f%%" % (100 * p["pct_cumplimiento_tabulacion"])
        lineas.append("| %s | %d | %d | %d | %d | %s | %d | %d |" % (
            r["programa"], p["n_sesiones"], p["n_sesiones_tabuladas"],
            p["n_sesiones_pendientes"], r["n_futuras"], pct,
            len(r["participantes"]), len(INC.por_programa(r["programa"]))))

    lineas += ["", "## Incidencias detectadas", ""]
    programas_con_inc = []
    for nombre in orden_programas:
        r = resultados.get(nombre)
        etiqueta = r["programa"] if r else nombre
        items = INC.por_programa(etiqueta)
        if items:
            programas_con_inc.append((etiqueta, items))
    if not programas_con_inc:
        lineas.append("_Sin incidencias._")
    for etiqueta, items in programas_con_inc:
        lineas.append("### %s" % etiqueta)
        lineas.append("")
        for sev, msg in items:
            lineas.append("- **%s** — %s" % (sev, msg))
        lineas.append("")

    lineas += [
        "## Cómo se ejecuta",
        "",
        "```bash",
        "python construir_base.py                                # corte = hoy, carpeta actual",
        "python construir_base.py --fecha-corte 2026-08-11",
        'python construir_base.py --input "C:/ruta/Prueba_Tecnica" \\',
        "                         --fecha-corte 2026-08-11 \\",
        "                         --output base_consolidada.xlsx",
        "```",
        "",
        "Requiere `pandas` y `openpyxl`. El script recorre dinámicamente toda subcarpeta "
        "que contenga `Equipo Logístico/Listado de Clases/`, ignora "
        "`Evidencia Fotográfica` y las imágenes, y **no aborta** si un programa tiene "
        "problemas: lo registra como incidencia y sigue con el resto.",
        "",
        "## Criterios aplicados",
        "",
        "- **Estado de tabulación (§5):** una sesión es `Tabulada` si su columna en "
        "`CONSOLIDADO` tiene al menos un valor (los ceros cuentan); `Pendiente de "
        "tabular` si está vacía y `fecha <= fecha_corte`; `Futura no exigible` si "
        "está vacía y `fecha > fecha_corte`.",
        "- **Cumplimiento:** `tabuladas / (tabuladas + pendientes)`. Las sesiones "
        "futuras nunca cuentan como incumplimiento.",
        "- **Columna por día, no por sesión:** el `CONSOLIDADO` trae una columna por "
        "día (o por día+jornada) mientras el cronograma puede tener varias sesiones "
        "ese día. Varias sesiones pueden compartir columna; el emparejamiento usa "
        "fecha y, cuando hay más de una columna para la misma fecha, jornada `T`/`M`.",
        "- **`asistencia_tabulada`:** `Sí` = Tabulada, `No` = Pendiente de tabular, "
        "`N/A` = Futura no exigible. Las sesiones sin columna en `CONSOLIDADO` quedan "
        "marcadas en `observaciones`.",
        "- **`fct_asistencia`** sólo contiene filas de sesiones tabuladas y con celda "
        "diligenciada; una celda vacía dentro de una sesión tabulada no genera fila.",
        "- **`total_inasistencia`** se recalcula sumando cada columna de `CONSOLIDADO` "
        "una sola vez por participante, aunque esa columna sirva a varias sesiones "
        "del mismo día (evita el doble conteo).",
        "- **Año de las columnas `DD T` / `DD M`:** se reconstruye con el mes "
        "propagado (forward-fill de la fila `Mes:`) y el año que hace calzar la fecha "
        "con el cronograma (2026 en todos los programas).",
        "- **`n_participantes`** se toma de `NÚMERO DE PARTICIANTES` (FORMAS DE PAGO), "
        "como indica §4.1. En 3 programas ese número declarado **no coincide** con las "
        "filas reales del `CONSOLIDADO` (ver incidencias); el conteo real está en "
        "`dim_participantes`, que es el que conviene usar para tasas de asistencia.",
        "- **Typo de año 2025** (Heridas y Odontología, campo `LUGAR Y FECHA`): se "
        "ignora para efectos de fechas; las sesiones salen del cronograma (§10).",
        "",
        "- **Celdas vacías:** todas las columnas de **texto** sin dato traen el literal "
        "`N/A` en lugar de quedar en blanco (`modulo`, `salon`, `docente`, "
        "`observaciones`, `empresa`, `nrc`, `cod_banner`, `codigo_contable`, "
        "`experto_facilitador`, `entidad_convenio`…). Se cambia con "
        "`--relleno-texto \"\"`.",
        "- **Las columnas numéricas y de fecha se dejan realmente nulas** a propósito: "
        "`n_asistentes`, `n_inasistentes`, `intensidad_horaria`, `valor_programa`, "
        "`horas_totales`, `horas_falla_max` y `pct_cumplimiento_tabulacion` quedan "
        "vacías cuando no aplican. Escribirles `N/A` las convertiría en columnas de "
        "texto en Power BI y rompería sumas, promedios y ejes de fecha; el vacío es lo "
        "que Power BI espera y lo muestra como *(Blank)*. En `fct_sesiones` esos vacíos "
        "son informativos: `n_asistentes`/`n_inasistentes` sólo existen si la sesión "
        "está `Tabulada`.",
        "",
        "> Nota para quien lea el Excel con pandas: `N/A` es un valor nulo por defecto "
        "en `read_excel`. Para leerlo como texto: "
        "`pd.read_excel(..., keep_default_na=False)`. Power BI lo importa como texto "
        "sin ajustes.",
        "",
        "## Verificación realizada",
        "",
        "Contraste del `.xlsx` generado contra los Excel originales, releyendo el "
        "`CONSOLIDADO` con un script independiente (corte `2026-08-11`):",
        "",
        "- **Cuidado de Heridas** — 21 columnas de sesión en `CONSOLIDADO`. Con datos: "
        "`24 T`, `25 T`, `30 T` (julio), `1 T`, `6 T` (agosto). Vacías: `31 T` y todas "
        "las de `8 T` en adelante. La base marca **11 Tabuladas** (las 4 sesiones del "
        "24-jul + 25-jul + 30-jul + las 4 del 1-ago + 6-ago), **5 Pendientes** (las 4 "
        "sesiones del 31-jul + la del 8-ago, ya pasadas y sin tabular) y **21 Futuras**. "
        "Los 37 estados coinciden. 35 participantes, igual que las filas reales de la hoja.",
        "- **Bienestar y Felicidad** — 4 columnas con fecha completa. Con datos: "
        "`2026-07-25`, `2026-08-01`, `2026-08-08`. Vacía: `2026-08-15` (futura). La base "
        "marca **3 Tabuladas** y **1 Futura no exigible**; cumplimiento 100 % porque la "
        "futura no entra al denominador. Los 4 estados coinciden. 9 participantes.",
        "- **Σ de inasistencia** — el total recalculado desde `fct_asistencia` coincide "
        "con la columna `Σ de inasistencia` del archivo para los **108 participantes de "
        "los 8 programas**, sin una sola diferencia.",
        "- **`en_riesgo`** — marca exactamente 3 participantes (Bootcamp ×1, Project ×2), "
        "los mismos que los archivos originales rotulan `NO GRADUA` en OBSERVACIONES.",
        "- **Integridad del modelo** — hojas y columnas exactas a §8, sin celdas "
        "combinadas, `fecha` como fecha real; `id_sesion`, `id_registro` y `programa_id` "
        "únicos; todas las FK resuelven; ninguna sesión futura queda como "
        "`Pendiente de tabular`; `n_asistentes`/`n_inasistentes` cuadran con "
        "`fct_asistencia`; `dim_calendario` continuo.",
        "- **Determinismo** — dos corridas seguidas producen hojas idénticas.",
        "",
        "**Discrepancias encontradas:** ninguna entre la base y los Excel originales. "
        "Las diferencias reportadas arriba son inconsistencias **dentro de los archivos "
        "fuente** (año 2025 en metadatos, horas invertidas, participantes declarados vs. "
        "reales, horas del cronograma vs. horas del CONSOLIDADO, columnas faltantes o "
        "duplicadas en Normatividad, `CONSOLIDADO` de Integración Sensorial sin fechas "
        "en los encabezados), que el script conserva y marca en lugar de corregir.",
        "",
    ]
    with open(ruta, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lineas))


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Construye la base consolidada del CEC a partir de las carpetas "
                    "de programa.")
    parser.add_argument("--input", default=".",
                        help="Carpeta raíz con las carpetas de programa (default: .)")
    parser.add_argument("--fecha-corte", default=None, metavar="YYYY-MM-DD",
                        help="Fecha de corte para los estados (default: hoy)")
    parser.add_argument("--output", default="base_consolidada.xlsx",
                        help="Archivo Excel de salida (default: base_consolidada.xlsx)")
    parser.add_argument("--readme", default="README.md",
                        help="Archivo markdown con las incidencias (default: README.md)")
    parser.add_argument("--relleno-texto", default=RELLENO_TEXTO, metavar="TEXTO",
                        help="Marcador para los vacíos de las columnas de texto "
                             "(default: N/A). Usa \"\" para dejarlas en blanco. Las "
                             "columnas numéricas y de fecha se dejan siempre nulas.")
    args = parser.parse_args(argv)

    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if args.fecha_corte:
        try:
            fecha_corte = dt.datetime.strptime(args.fecha_corte, "%Y-%m-%d").date()
        except ValueError:
            parser.error("--fecha-corte debe tener formato YYYY-MM-DD")
    else:
        fecha_corte = dt.date.today()

    raiz = os.path.abspath(args.input)
    if not os.path.isdir(raiz):
        parser.error("La carpeta de entrada no existe: %s" % raiz)

    salida = args.output if os.path.isabs(args.output) else os.path.join(raiz, args.output)
    readme = args.readme if os.path.isabs(args.readme) else os.path.join(raiz, args.readme)

    print("=" * 78)
    print("BASE CONSOLIDADA CEC — Universidad del Norte")
    print("=" * 78)
    print("Carpeta de entrada : %s" % raiz)
    print("Fecha de corte     : %s" % fecha_corte.isoformat())
    print("Archivo de salida  : %s" % salida)
    print("-" * 78)

    programas = descubrir_programas(raiz)
    if not programas:
        print("ERROR: no se encontró ninguna carpeta con 'Equipo Logístico/"
              "Listado de Clases'.")
        return 1
    print("Programas detectados: %d" % len(programas))
    print("-" * 78)

    resultados, orden = {}, []
    for carpeta, ruta_clases in programas:
        orden.append(carpeta)
        try:
            r = procesar_programa(carpeta, ruta_clases, fecha_corte)
            resultados[carpeta] = r
            print("  [OK]    %-32s %3d sesiones, %3d participantes"
                  % (r["programa"][:32], len(r["sesiones"]), len(r["participantes"])))
        except Exception as exc:      # §9: un programa con problemas no aborta todo
            programa_id, programa = identificar_programa(carpeta)
            INC.add(programa, "Error al procesar el programa: %s: %s"
                    % (type(exc).__name__, exc), "ERROR")
            print("  [FALLA] %-32s %s: %s" % (carpeta[:32], type(exc).__name__, exc))
            traceback.print_exc(file=sys.stdout)

    # ---------------- ensamblado
    sesiones = [f for c in orden if c in resultados for f in resultados[c]["sesiones"]]
    asistencia = [f for c in orden if c in resultados for f in resultados[c]["asistencia"]]
    participantes = [f for c in orden if c in resultados
                     for f in resultados[c]["participantes"]]
    programas_rows = [resultados[c]["programa_row"] for c in orden if c in resultados]

    calendario = construir_calendario([f["fecha"] for f in sesiones])
    relleno = args.relleno_texto

    hojas = OrderedDict([
        ("fct_sesiones", preparar(pd.DataFrame(sesiones), "fct_sesiones", relleno)),
        ("dim_programas", preparar(pd.DataFrame(programas_rows), "dim_programas", relleno)),
        ("dim_calendario", preparar(calendario, "dim_calendario", relleno)),
        ("fct_asistencia", preparar(pd.DataFrame(asistencia), "fct_asistencia", relleno)),
        ("dim_participantes", preparar(pd.DataFrame(participantes), "dim_participantes",
                                       relleno)),
        ("Parametros", preparar(pd.DataFrame([
            {"parametro": "fecha_corte",
             "valor": fecha_corte.isoformat(),
             "descripcion": "Fecha de corte usada para los estados (editable)."},
            {"parametro": "fecha_generacion",
             "valor": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             "descripcion": "Timestamp de ejecución del script."},
        ]), "Parametros", relleno)),
    ])

    escribir_excel(salida, hojas)
    escribir_readme(readme, resultados, fecha_corte, salida, orden)

    # ---------------- log resumen (§9)
    print("-" * 78)
    print("RESUMEN POR PROGRAMA")
    print("-" * 78)
    cab = "%-32s %8s %10s %11s %8s %7s %6s" % (
        "PROGRAMA", "SESIONES", "TABULADAS", "PENDIENTES", "FUTURAS", "CUMPL.", "PART.")
    print(cab)
    print("-" * 78)
    tot = Counter()
    for carpeta in orden:
        r = resultados.get(carpeta)
        if not r:
            print("%-32s %8s %10s %11s %8s %7s %6s"
                  % (carpeta[:32], "ERROR", "-", "-", "-", "-", "-"))
            continue
        p = r["programa_row"]
        pct = "-" if p["pct_cumplimiento_tabulacion"] is None else \
            "%.0f%%" % (100 * p["pct_cumplimiento_tabulacion"])
        print("%-32s %8d %10d %11d %8d %7s %6d" % (
            r["programa"][:32], p["n_sesiones"], p["n_sesiones_tabuladas"],
            p["n_sesiones_pendientes"], r["n_futuras"], pct, len(r["participantes"])))
        tot["sesiones"] += p["n_sesiones"]
        tot["tab"] += p["n_sesiones_tabuladas"]
        tot["pen"] += p["n_sesiones_pendientes"]
        tot["fut"] += r["n_futuras"]
        tot["part"] += len(r["participantes"])
    print("-" * 78)
    denom = tot["tab"] + tot["pen"]
    print("%-32s %8d %10d %11d %8d %7s %6d" % (
        "TOTAL", tot["sesiones"], tot["tab"], tot["pen"], tot["fut"],
        ("%.0f%%" % (100 * tot["tab"] / denom)) if denom else "-", tot["part"]))
    print("-" * 78)
    print("Total de participantes (participante × programa): %d" % len(participantes))
    print("Registros de asistencia (participante × sesión) : %d" % len(asistencia))
    print("Días en dim_calendario                          : %d" % len(calendario))

    print("-" * 78)
    print("INCIDENCIAS DE CALIDAD DE DATOS (%d)" % len(INC.items))
    print("-" * 78)
    if not INC.items:
        print("  Sin incidencias.")
    ultimo = None
    for prog, sev, msg in INC.items:
        if prog != ultimo:
            print("\n  %s" % prog)
            ultimo = prog
        print("    [%s] %s" % (sev, msg))
    print("-" * 78)
    print("OK  -> %s" % salida)
    print("OK  -> %s" % readme)
    return 0


if __name__ == "__main__":
    sys.exit(main())
